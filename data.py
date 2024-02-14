import openpyxl

class classData:
    def __init__(self,driver):
        self.driver = driver

    def getData():
        excel = openpyxl.load_workbook("veri.xlsx")
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            senaryoturu = sheet.cell(i,1).value
            kelime = sheet.cell(i,2).value
            gerceklesen_sonuc = sheet.cell(i,3).value
            data.append((senaryoturu,kelime,gerceklesen_sonuc))
        return data